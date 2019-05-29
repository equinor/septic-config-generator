from jinja2 import Environment, FileSystemLoader
from scg.config_parser import parse_config
import os
from openpyxl import load_workbook

def read_source(source, root):
    wb = load_workbook(os.path.join(root, source['filename']), True)
    sheet = wb[source['sheet']]
    rows = sheet.max_row
    cols = sheet.max_column
    headers = dict((i, sheet.cell(row=1, column=i).value) for i in range(1, cols+1))
    ret = dict()
    for i in range(2, rows + 1):
        temp = dict()
        for j in range(2, cols + 1):
            temp[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            ret[sheet.cell(row=i, column=1).value] = temp
    return ret
    #def item(i, j):
    #    return (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=j).value)
    #return (dict(item(i, j) for j in range(1, cols + 1 )) for i in range(2, rows + 1))


if __name__ == '__main__':
    cfg = parse_config('johan_sverdrup.yaml').data

    template_paths = [os.path.join(cfg['path']['root'], p) for p in cfg['path']['templatepaths']]
    env = Environment(
        loader=FileSystemLoader(searchpath=[os.path.join(cfg['path']['root'], p) for p in cfg['path']['templatepaths']])
    )
    env.variable_start_string = ''
    env.variable_end_string = ''
    sources = dict()
    for source in cfg['sources']:
        s = read_source(source, cfg['path']['root'])
        sources[source['id']] = s

    with open(cfg['output'], 'w') as f:
        for template in cfg['layout']:
            temp = env.get_template(template['name'])
            if not 'source' in template:
                print(temp.render({}))
                continue
            items = []
            if 'include' in template:
                items = template['include']
            else:
                items = list(sources[template['source']].keys())
                if 'exclude' in template:
                    items = [x for x in items if x not in template['exclude']]

            for row, values in sources[template['source']].items():
                if row in items:
                    print(temp.render(values))
                    f.write(temp.render(values))


