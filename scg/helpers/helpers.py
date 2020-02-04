import os
import sys
import shutil
from openpyxl import load_workbook
import logging
import difflib

logger = logging.getLogger("scg." + __name__)


def diff_cnfgs(original_config, new_config):
    orig = open(original_config).readlines()
    new = open(new_config).readlines()
    return difflib.unified_diff(orig, new, fromfile=original_config, tofile=new_config)


def diff_backup_and_replace(original, new, verify=True):
    if os.path.exists(original):
        backup = original + ".bak"
        if verify:
            origtxt = open(original).readlines()
            newtxt = open(new).readlines()
            diff = difflib.unified_diff(origtxt, newtxt, fromfile=original, tofile=new)
            txt = [line for line in diff]
            if len(txt) > 0:
                print("".join(txt))
                q = input("Replace original? [Y]es or [N]o: ")
                if len(q) > 0 and q[0].lower() == "y":
                    if os.path.isfile(backup):
                        os.remove(backup)
                    shutil.move(original, backup)
                    shutil.move(new, original)
                else:
                    os.remove(new)
            else:
                logger.info("No change. Keeping original.")
                os.remove(new)
        else:
            shutil.move(original, backup)
            shutil.move(new, original)
    else:
        shutil.move(new, original)


def read_source(source, root):
    if not source["filename"].endswith(".xlsx"):
        logger.error(
            f"Source files need to be xlsx. Please check source '{source['id']}'"
        )
        sys.exit()
    filename = source["filename"]
    try:
        wb = load_workbook(os.path.join(root, filename), read_only=True, data_only=True)
    except:
        logger.error(
            f"Unable to read '{source['filename']}' which belongs to source '{source['id']}'"
        )
        sys.exit(1)

    sheet = wb[source["sheet"]]
    cols = sheet.max_column
    ret = dict()
    for rownum, row in enumerate(sheet, 1):
        if rownum == 1:
            continue
        if all(c.value is None for c in row):
            break
        temp = dict()
        for j in range(2, cols + 1):
            if sheet.cell(row=1, column=j).value is None:
                break
            temp[sheet.cell(row=1, column=j).value] = str(
                sheet.cell(row=rownum, column=j).value
            )
            ret[sheet.cell(row=rownum, column=1).value] = temp
    return ret


def get_all_source_data(sources, path):
    res = dict()
    for source in sources:
        s = read_source(source, path)
        res[source["id"]] = s
    return res
